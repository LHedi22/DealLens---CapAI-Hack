from openpyxl import load_workbook


def parse_xlsx(file_path: str) -> str:
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        parts = []
        for sheet in wb.worksheets:
            parts.append(f"Sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(v) for v in row if v is not None)
                if row_text.strip():
                    parts.append(row_text)
        return "\n".join(parts)
    except Exception:
        return ""
